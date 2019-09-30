#! /usr/bin/env python3

import sys
import openpyxl
from PIL import Image

class Generator():
    """
    """
    def __init__(self, image):
        """
        """
        self.image = Image.open(image)
        self.image_width = self.image.size[0]
        self.image_height = self.image.size[1]
        self.image_pix = self.image.load()
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

    def rgb2hex(self, rgb):
        """
        """
        hex_value = '#%02x%02x%02x' % rgb
        return hex_value

    def colnum_string(self, n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def draw(self):
        """
        """
        for row in range(0, self.image_height):
            for col in range(0, self.image_width):
                color_rgb = self.image_pix[col, row]
                color_hex = self.rgb2hex(color_rgb)[1:]
                cell = self.worksheet.cell(row=row+1, column=col+1)
                self.worksheet.row_dimensions[row+1].height = 5
                self.worksheet.column_dimensions[self.colnum_string(col+1)].width = 5
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor=color_hex)
        self.workbook.save("test.xlsx")



if __name__ == "__main__":
    image = sys.argv[1]

    GEN = Generator(image)
    GEN.draw()